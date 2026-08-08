import os
import re
import requests
from openpyxl import load_workbook

ARQUIVO = "casas_oficiais.xlsx"
PASTA = "assets/bets"

URL_PLANILHA = (
    "https://www.gov.br/fazenda/pt-br/composicao/orgaos/"
    "secretaria-de-premios-e-apostas/lista-de-empresas/"
    "planilha-de-autorizacoes.xlsx"
)

os.makedirs(PASTA, exist_ok=True)

headers = {
    "User-Agent": "Mozilla/5.0"
}

print("Baixando planilha oficial do Ministério da Fazenda...")

r = requests.get(URL_PLANILHA, headers=headers, timeout=60)
r.raise_for_status()

with open(ARQUIVO, "wb") as f:
    f.write(r.content)

print("✓ Planilha baixada")

wb = load_workbook(ARQUIVO, data_only=True)

dominios = set()

for ws in wb.worksheets:
    for row in ws.iter_rows(values_only=True):
        for valor in row:
            if valor is None:
                continue

            texto = str(valor).lower()

            encontrados = re.findall(
                r"[a-z0-9][a-z0-9.-]*\.bet\.br",
                texto
            )

            dominios.update(encontrados)

dominios = sorted(dominios)

print()
print(f"Encontrados {len(dominios)} domínios .bet.br")
print()

for i, dominio in enumerate(dominios, 1):

    nome = dominio.replace(".bet.br", "")
    nome = re.sub(r"[^a-z0-9_-]", "_", nome)

    url_logo = (
        "https://www.google.com/s2/favicons"
        f"?domain=https://{dominio}&sz=256"
    )

    try:
        resp = requests.get(
            url_logo,
            headers=headers,
            timeout=20
        )

        if resp.status_code == 200 and len(resp.content) > 100:

            caminho = os.path.join(
                PASTA,
                f"{nome}.png"
            )

            with open(caminho, "wb") as f:
                f.write(resp.content)

            print(
                f"[{i}/{len(dominios)}] ✓ "
                f"{dominio}"
            )

        else:
            print(
                f"[{i}/{len(dominios)}] ✗ "
                f"{dominio}"
            )

    except Exception as erro:
        print(
            f"[{i}/{len(dominios)}] ✗ "
            f"{dominio}: {erro}"
        )

with open("dominios_bets.txt", "w") as f:
    for dominio in dominios:
        f.write(dominio + "\n")

print()
print("==============================")
print("FINALIZADO")
print("==============================")
print(f"Total de casas/domínios: {len(dominios)}")
print(f"Logos: {PASTA}")
print("Lista: dominios_bets.txt")
